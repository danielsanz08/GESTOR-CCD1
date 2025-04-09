from django.shortcuts import render
from papeleria.models import Articulo
from libreria.models import CustomUser
from papeleria.forms import LoginForm, ArticuloForm, ArticuloEditForm
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.template.loader import get_template
from django.utils.timezone import now
from django.http import HttpResponse
from xhtml2pdf import pisa
from io import BytesIO
# Create your views here.
def index_pap(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
    ]
    return render(request, 'index_pap/index_pap.html', {'breadcrumbs': breadcrumbs})
def login_papeleria(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, email=email, password=password)

            if user is not None:
                if user.module == 'Papeleria':  # Verifica que el usuario pertenece a Papelería
                    login(request, user)
                    messages.success(request, "Sesión iniciada correctamente en Papelería.")
                    return redirect('papeleria:index_pap')  # Redirige a la página de inicio de Papelería
                else:
                    messages.error(request, "No tienes acceso a este módulo.")
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = LoginForm()

    return render(request, 'login_pap/login_pap.html', {'form': form})

# CERRAR SESIÓN
def logout_view(request):
    logout(request)
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect(reverse('libreria:inicio'))
User = get_user_model()


def crear_articulo(request):
    if request.method == 'POST':
        form = ArticuloForm(request.POST)
        if form.is_valid():  # Corrección aquí
            articulo = form.save(commit=False)
            articulo.registrado_por = request.user
            articulo.save()
            return redirect('papeleria:listar_articulo')  # Redirección correcta
    else:
        form = ArticuloForm()
    
    return render(request, 'articulo/crear_articulo.html', {'form': form})
def editar_articulo(request, articulo_id):
    articulo = get_object_or_404(Articulo, id=articulo_id)
    if request.method== 'POST':
        form = ArticuloEditForm(request.POST, instance=articulo)
        if form.is_valid():
            form.save()
            return redirect('papeleria:listar_articulo')
    else:
            form = ArticuloEditForm(instance=articulo)
    return render(request, 'articulo/editar_articulo.html', {'form': form, 'articulo': articulo})


def listar_articulo(request):
    query = request.GET.get('q', '').strip()
    articulos = Articulo.objects.select_related('registrado_por').all()


    if query:
        articulos = articulos.filter(
        Q(nombre__icontains=query) |
        Q(marca__icontains=query) |
        Q(observacion__icontains=query) |
        Q(tipo__icontains=query) |
        Q(precio__icontains=query) |
        Q(cantidad__icontains=query) |
        Q(registrado_por__username__icontains=query) |
        Q(fecha_registro__icontains=query)
    )

    paginator = Paginator(articulos, 5)
    page_number = request.GET.get('page')
    articulos = paginator.get_page(page_number)

    return render(request, 'articulo/listar_articulo.html', {'articulos': articulos})
def buscar_articulo(request):
    query = request.GET.get('q', '').strip()
    if query:
        articulos = Articulo.objects.filter(nombre__icontains=query).values(
            'id', 'nombre', 'marca', 'observacion', 'precio', 'registrado_por', 'fecha_registro'
        )
        return JsonResponse(list(articulos), safe=False)
    return JsonResponse([], safe=False)

def eliminar_articulo(request, id):
    articulo = Articulo.objects.get(id=id)
    if request.method == 'POST':
        articulo.delete()
        messages.success(request, "Artículo eliminado correctamente.")
    return redirect('papeleria:listar_articulo')


def validar_datos(request):
    email = request.GET.get('email', None)
    
    errores = {}

    # Validar correo electrónico
    if email and CustomUser.objects.filter(email=email).exists():
        errores['email'] = 'El email ya está en uso.'

    # Retornar los errores (si los hay) o una respuesta de validación exitosa
    return JsonResponse(errores if errores else {'valid': True})
def verificar_nombre_articulo(request):
    nombre = request.GET.get('nombre', '')
    existe = Articulo.objects.filter(nombre=nombre).exists()
    return JsonResponse({'existe': existe})

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from django.contrib.staticfiles import finders
from reportlab.lib import colors  
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import inch
from datetime import datetime


def reporte_articulo_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    margin = 40
    table_width = width - 2 * margin
    y_position = height - margin - 110

    # Marca de agua
    watermark_path = finders.find('imagen/LOGO.png')
    p.saveState()
    p.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
    p.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
    p.drawImage(watermark_path, x=(width - 600) / 2, y=(height - 600) / 2, width=600, height=600, mask='auto')
    p.restoreState()

    # Título
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2, height - margin - 30, "REPORTE DE ARTÍCULOS")

    # Encabezado de empresa
    tabla_encabezado = Table([
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "contacto@gestorccd.com", "Teléfono: (123) 456-7890"],
    ], colWidths=[180, 180, 180, 180])

    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    tabla_encabezado.wrapOn(p, table_width, height)
    tabla_encabezado.drawOn(p, margin, y_position)

    # Tabla del usuario autenticado (justo debajo del encabezado)
    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        usuario.cargo
    ])

    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    table_usuario.wrapOn(p, table_width, height)

    # Calcular la altura ocupada por el encabezado para colocar al usuario pegado debajo
    alto_encabezado = len(tabla_encabezado._cellvalues) * 18
    y_tabla_usuario = y_position - alto_encabezado -12 # Espacio adicional de 20 unidades
    table_usuario.drawOn(p, margin, y_tabla_usuario)

    # Tabla de artículos (pegada debajo del usuario)
    data_articulos = [["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]]
    articulos = Articulo.objects.all()
    for articulo in articulos:
        data_articulos.append([
            articulo.id,
            articulo.nombre,
            articulo.marca,
            articulo.tipo,
            articulo.precio,
            articulo.cantidad,
            articulo.observacion
        ])

    table = Table(data_articulos, colWidths=[70, 100, 100, 90, 90, 90, 180])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    table.wrapOn(p, table_width, height)

    alto_usuario = len(data_usuario) * 18
    y_tabla_articulos = y_tabla_usuario - alto_usuario-45
    table.drawOn(p, margin, y_tabla_articulos)

    # Finalizar PDF
    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_articulos.pdf"'
    return response
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def reporte_articulo_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Artículos"

    # Ajustar el ancho de columnas A y B para que quepa el logo
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    # Ajustar la altura de la fila 1 para el logo
    ws.row_dimensions[1].height = 90

    # Agregar el logo
    logo_path = finders.find('imagen/logo.png')
    img = Image(logo_path)
    img.height = 80
    img.width = 200
    ws.add_image(img, 'A1')

    # Fusionar celdas A1 y B1 para que el logo ocupe espacio visualmente
    ws.merge_cells('A1:B1')

    # Título principal
    ws.merge_cells('C1:F1')
    ws['C1'] = "GESTOR CCD"
    ws['C1'].font = Font(size=24, bold=True)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo
    ws.merge_cells('A2:G2')
    ws['A2'] = "Listado de Artículos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados de la tabla
    headers = ["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]
    ws.append(headers)

    # Estilo para los encabezados
    header_fill = PatternFill(start_color="01AB7B", end_color="01AB7B", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Bordes
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Agregar datos
    articulos = Articulo.objects.all()
    for articulo in articulos:
        ws.append([
            articulo.id,
            articulo.nombre,
            articulo.marca,
            articulo.tipo,
            articulo.precio,
            articulo.cantidad,
            articulo.observacion,
        ])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30

    # Aplicar estilo a celdas de datos
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Altura de filas para los títulos
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Preparar la respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_articulos.xlsx"'
    wb.save(response)
    return response